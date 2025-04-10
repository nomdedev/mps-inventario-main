import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from fpdf import FPDF
from datetime import datetime

EXPORT_FOLDER = "exports"

def _crear_encabezado(sheet, headers):
    """
    Crea un encabezado en la hoja de Excel con estilos básicos.
    :param sheet: Hoja de Excel.
    :param headers: Lista de encabezados.
    """
    for col_num, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

def exportar_materiales_obra(obra_id, materiales):
    """
    Exporta los materiales de una obra a un archivo Excel.
    :param obra_id: ID de la obra.
    :param materiales: Lista de diccionarios con los datos de los materiales.
    :return: Ruta del archivo generado.
    """
    if not os.path.exists(EXPORT_FOLDER):
        os.makedirs(EXPORT_FOLDER)

    filename = os.path.join(EXPORT_FOLDER, f"obra_{obra_id}_materiales.xlsx")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Materiales"

    # Crear encabezado
    headers = ["Código", "Descripción", "Cantidad", "Fecha de Apartado", "Usuario"]
    _crear_encabezado(sheet, headers)

    # Agregar datos
    for row_num, material in enumerate(materiales, start=2):
        sheet.cell(row=row_num, column=1, value=material["codigo"])
        sheet.cell(row=row_num, column=2, value=material["descripcion"])
        sheet.cell(row=row_num, column=3, value=material["cantidad"])
        sheet.cell(row=row_num, column=4, value=material["fecha_apartado"])
        sheet.cell(row=row_num, column=5, value=material["usuario"])

    workbook.save(filename)
    return filename

def exportar_pedido(pedido_id, items):
    """
    Exporta los ítems de un pedido a un archivo Excel.
    :param pedido_id: ID del pedido.
    :param items: Lista de diccionarios con los datos de los ítems.
    :return: Ruta del archivo generado.
    """
    if not os.path.exists(EXPORT_FOLDER):
        os.makedirs(EXPORT_FOLDER)

    filename = os.path.join(EXPORT_FOLDER, f"pedido_{pedido_id}.xlsx")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Pedido"

    # Crear encabezado
    headers = ["Material ID", "Descripción", "Cantidad", "Estado"]
    _crear_encabezado(sheet, headers)

    # Agregar datos
    for row_num, item in enumerate(items, start=2):
        sheet.cell(row=row_num, column=1, value=item["material_id"])
        sheet.cell(row=row_num, column=2, value=item["descripcion"])
        sheet.cell(row=row_num, column=3, value=item["cantidad"])
        sheet.cell(row=row_num, column=4, value=item["estado"])

    workbook.save(filename)
    return filename

def exportar_movimientos(material_id, movimientos):
    """
    Exporta los movimientos de un material a un archivo Excel.
    :param material_id: ID del material.
    :param movimientos: Lista de diccionarios con los datos de los movimientos.
    :return: Ruta del archivo generado.
    """
    if not os.path.exists(EXPORT_FOLDER):
        os.makedirs(EXPORT_FOLDER)

    filename = os.path.join(EXPORT_FOLDER, f"movimientos_{material_id}.xlsx")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Movimientos"

    # Crear encabezado
    headers = ["Fecha", "Tipo", "Cantidad", "Usuario", "Obra ID", "Estado"]
    _crear_encabezado(sheet, headers)

    # Agregar datos
    for row_num, movimiento in enumerate(movimientos, start=2):
        sheet.cell(row=row_num, column=1, value=movimiento["fecha"])
        sheet.cell(row=row_num, column=2, value=movimiento["tipo"])
        sheet.cell(row=row_num, column=3, value=movimiento["cantidad"])
        sheet.cell(row=row_num, column=4, value=movimiento["usuario"])
        sheet.cell(row=row_num, column=5, value=movimiento["obra_id"])
        sheet.cell(row=row_num, column=6, value=movimiento["estado"])

    workbook.save(filename)
    return filename

def exportar_checklist_final(obra_id, materiales_entregados, faltantes):
    """
    Exporta un checklist final de una obra a un archivo PDF.
    :param obra_id: ID de la obra.
    :param materiales_entregados: Lista de materiales entregados.
    :param faltantes: Lista de materiales faltantes.
    :return: Ruta del archivo generado.
    """
    if not os.path.exists(EXPORT_FOLDER):
        os.makedirs(EXPORT_FOLDER)

    filename = os.path.join(EXPORT_FOLDER, f"finalizacion_obra_{obra_id}.pdf")
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Arial", size=12)

    # Título
    pdf.set_font("Arial", style="B", size=16)
    pdf.cell(0, 10, f"Checklist Final - Obra ID: {obra_id}", ln=True, align="C")
    pdf.ln(10)

    # Materiales entregados
    pdf.set_font("Arial", style="B", size=14)
    pdf.cell(0, 10, "Materiales Entregados:", ln=True)
    pdf.set_font("Arial", size=12)
    if materiales_entregados:
        for material in materiales_entregados:
            pdf.cell(0, 10, f"- {material['descripcion']} (Cantidad: {material['cantidad']})", ln=True)
    else:
        pdf.cell(0, 10, "No se entregaron materiales.", ln=True)
    pdf.ln(10)

    # Materiales faltantes
    pdf.set_font("Arial", style="B", size=14)
    pdf.cell(0, 10, "Materiales Faltantes:", ln=True)
    pdf.set_font("Arial", size=12)
    if faltantes:
        for material in faltantes:
            pdf.cell(0, 10, f"- {material['descripcion']} (Cantidad: {material['cantidad']})", ln=True)
    else:
        pdf.cell(0, 10, "No hay materiales faltantes.", ln=True)
    pdf.ln(10)

    # Fecha y firma
    pdf.set_font("Arial", style="B", size=14)
    pdf.cell(0, 10, "Fecha de Cierre:", ln=True)
    pdf.set_font("Arial", size=12)
    pdf.cell(0, 10, "__________________________", ln=True)
    pdf.ln(10)

    pdf.set_font("Arial", style="B", size=14)
    pdf.cell(0, 10, "Firma Responsable:", ln=True)
    pdf.set_font("Arial", size=12)
    pdf.cell(0, 10, "__________________________", ln=True)

    pdf.output(filename)
    return filename

def exportar_dashboard(usuario_actual, datos):
    """
    Exporta el resumen del dashboard a un archivo PDF.
    :param usuario_actual: Usuario que genera el reporte.
    :param datos: Diccionario con los datos del dashboard.
    :return: Ruta del archivo generado.
    """
    if not os.path.exists(EXPORT_FOLDER):
        os.makedirs(EXPORT_FOLDER)

    filename = os.path.join(EXPORT_FOLDER, f"dashboard_resumen_{usuario_actual}.pdf")
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Arial", size=12)

    # Título
    pdf.set_font("Arial", style="B", size=16)
    pdf.cell(0, 10, "Resumen del Dashboard", ln=True, align="C")
    pdf.ln(10)

    # Usuario y fecha
    pdf.set_font("Arial", size=12)
    pdf.cell(0, 10, f"Generado por: {usuario_actual}", ln=True)
    pdf.cell(0, 10, f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", ln=True)
    pdf.ln(10)

    # KPIs
    pdf.set_font("Arial", style="B", size=14)
    pdf.cell(0, 10, "KPIs:", ln=True)
    pdf.set_font("Arial", size=12)
    for key, value in datos["totales"].items():
        pdf.cell(0, 10, f"- {key.replace('_', ' ').capitalize()}: {value}", ln=True)
    pdf.ln(10)

    # Entregas recientes
    pdf.set_font("Arial", style="B", size=14)
    pdf.cell(0, 10, "Entregas Recientes:", ln=True)
    pdf.set_font("Arial", size=12)
    for entrega in datos["entregas"]:
        pdf.cell(0, 10, f"- Obra: {entrega['obra']}, Usuario: {entrega['usuario']}, Fecha: {entrega['fecha']}", ln=True)

    pdf.output(filename)
    return filename
